# 饿了么指定地点下，各个商家的信息

#益园文创基地
#https://www.ele.me/restapi/shopping/restaurants?extras[]=activities&geohash=wx4em10hdzmd&latitude=39.952594&limit=24&longitude=116.235477&offset=0&terminal=web

#菜单
# https://www.ele.me/restapi/shopping/v2/menu?restaurant_id=163186695&terminal=web

# https://mainsite-restapi.ele.me/v2/pois?
# extras%5B%5D=count&geohash=wx4g0bmjetr7&keyword=%E6%9C%9D%E9%98%B3&limit=20&type=nearby
import urllib.request
import os
import json
import time
from openpyxl import Workbook, load_workbook

targetDir = "C:\\Users\\dell\\Desktop\\temp\\爬虫饿了么"  # 文件保存路径


def excelName(name):  # 根据日期生成文件
    if not os.path.isdir(targetDir):
        os.mkdir(targetDir)
    excelName = name + str(time.strftime("%Y-%m-%d") + ".xlsx")
    completePath = targetDir + "\\" + excelName
    return completePath


def reqsetting():  # 初始化url请求，需要实时替换的是extral  和  header里的referer
    weburl = "https://mainsite-restapi.ele.me/shopping/restaurants?"
    extra1 = "extras%5B%5D=activities&geohash=wx4g56v1d2m&latitude=39.91771&limit=24&longitude=116.51698&offset=0&terminal=web"
    webheaders = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        "Cookie": "ubt_ssid=aljer72855q5kxhxn12ba0tylnxawxa8_2018-09-28; _utrace=908d16f73f446241d239a898c44f6b9d_2018-09-28; track_id=1538145248|77cf24f83819face4a33522503665aa609444b047f3ab200e4|4850253820df707e9623e2ad27a6a579; USERID=492048802; SID=2ZvKrBUPkVF3v1IdtepjciO3hs4jpD0QJztA",
        "Host": "mainsite-restapi.ele.me",
        "Origin": "https://www.ele.me",
        # "Referer":"https://www.ele.me/place/wx4g56v1d2m",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36"
    }
    req = urllib.request.Request(url=weburl, headers=webheaders)
    return req


def write2Excel(jsondata, title):  # 根据不同的商圈地点写入数据，每个商圈地点占用excel 的一个sheet
    fileName = excelName(title)
    if (os.path.exists(fileName)):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()
    if (wb.__contains__(title)):
        ws = wb[title]
        # ws.append([])
    else:
        ws = wb.create_sheet(title)
        ws.column_dimensions["A"].width = 10.0
        ws.column_dimensions["B"].width = 30.0
        ws.column_dimensions["C"].width = 10.0
        ws.column_dimensions["D"].width = 10.0
        ws.column_dimensions["E"].width = 10.0
        ws.column_dimensions["F"].width = 10.0
        ws.column_dimensions["G"].width = 20.0
        ws.column_dimensions["H"].width = 10.0
        ws.column_dimensions["I"].width = 10.0
        ws.column_dimensions["J"].width = 30.0
        ws.column_dimensions["K"].width = 10.0
        ws.column_dimensions["L"].width = 40.0
        ws.column_dimensions["M"].width = 10.0

        ws.append(["店铺ID", "店铺名称", "月销量", "起送价", "配送费", "平均送达速度", "营业时长", "评分", "评价数", "活动", "是否新开店", "店铺地址", "距离"])

    for i in range(len(jsondata)):
        attribute = ""
        row = jsondata[i]
        # print(type(row))
        for activitie in row["activities"]:
            if "type" in activitie and activitie['type'] is 102:
                attribute = attribute + activitie["tips"]
            pass
        print(row)
        ws.append(
            [row["id"], row["name"], row["recent_order_num"], row["float_minimum_order_amount"], row["float_delivery_fee"], row["order_lead_time"], str(row["opening_hours"]), row["rating"], row["rating_count"], str(attribute), row["is_new"], row["address"], row["distance"]])

    wb.save(fileName)

def readKeyWordFromExcel():
    req = reqsetting()
    req.add_header("Refer", "https://www.ele.me/place/wx4em10hdzmd?latitude=39.952594&longitude=116.235477")  # 修改请求头的refer
    newUrl = req.get_full_url()
    # wwc2kguv86xu?latitude=38.034253&longitude=114.473097
    # / wwc2nz95sv23?latitude = 38.010578 & longitude = 114.556021 杨麻子大饼  wwc82028ytkf?latitude=38.014226&longitude=114.610347
    params = {
        "extras[]": "activities",
        "geohash": "wwc82028ytkf",
        "latitude": "38.014226",
        "longitude": "114.610347",
        "terminal": "web",
        "limit": 30,
        "offset":90
    }
    params = urllib.parse.urlencode(params)  # 请求参数编码
    req.full_url = newUrl + params  # 重新生成url请求
    webpage = urllib.request.urlopen(req)
    contentBytes = webpage.read().decode("utf-8")

    jsondata = json.loads(contentBytes)
    write2Excel(jsondata, "遛鱼巫山烤全鱼")  # 将请求数据写入excel中



if __name__ == '__main__':  # 程序运行入口
    offset = 0;
    readKeyWordFromExcel()
