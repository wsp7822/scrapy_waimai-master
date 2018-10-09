# 爬饿了么外卖数据--区域集合

# https://mainsite-restapi.ele.me/v2/pois?
# extras%5B%5D=count&geohash=wx4g0bmjetr7&keyword=%E6%9C%9D%E9%98%B3&limit=20&type=nearby
import urllib.request
import os
import json
from openpyxl import Workbook
from openpyxl import load_workbook

keywordExcel = "C:\\Users\\dell\\Desktop\\temp\\keyword.xlsx"  # 关键字检索外卖地点保存路径

keywords = ["朝阳", "奥体"]  # 关键字集合


def reqsetting():  # 首先构造请求头headers，url目前暂时保存根路径

    weburl = "https://mainsite-restapi.ele.me/v2/pois?"
    # extra1="extras%5B%5D=count&geohash=wx4g0bmjetr7&keyword=%E6%9C%9D%E9%98%B3&limit=20&type=nearby"
    webheaders = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.8",
        "Connection": "keep-alive",
        "Cookie": "ubt_ssid=plds7ye19rj2rghg3oaar8hkt89yy7f1_2017-02-07; _utrace=ac9073c509bedb74b28a1482bd95a9d8_2017-02-07",
        "Host": "mainsite-restapi.ele.me",
        "Origin": "https://www.ele.me",
        "Referer": "https://www.ele.me/place/wx4g4h5shqf",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.75 Safari/537.36"
    }
    req = urllib.request.Request(url=weburl, headers=webheaders)

    return req


def write2Excel(jsondata, title):  # 根据不同的关键字将数据写入到excel中
    fileName = keywordExcel
    if (os.path.exists(fileName)):
        wb = load_workbook(fileName)
    else:
        wb = Workbook()

    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 10.0
    ws.append(["ID", "城市", "geohash", "名称", "地址", "商家总数", "经度", "纬度", "request_id", "short_address"])
    ws.column_dimensions["A"].width = 30.0
    ws.column_dimensions["B"].width = 10.0
    ws.column_dimensions["C"].width = 18.0
    ws.column_dimensions["D"].width = 20.0
    ws.column_dimensions["E"].width = 50.0
    ws.column_dimensions["F"].width = 10.0
    ws.column_dimensions["G"].width = 10.0
    ws.column_dimensions["H"].width = 10.0
    ws.column_dimensions["I"].width = 25.0
    ws.column_dimensions["J"].width = 40.0

    for i in range(len(jsondata)):
        row = jsondata[i]

        ws.append([row["id"], row["city"], row["geohash"], row["name"], row["address"], row["count"],
                   row["longitude"], row["latitude"], row["request_id"], row["short_address"]])
    wb.save(fileName)


if __name__ == '__main__':  # 程序运行入口

    if (os.path.exists(keywordExcel)):
        os.remove(keywordExcel)
    req = reqsetting()
    newUrl = req.get_full_url()
    for keyword in keywords:  # 遍历关键字集合，构造不同的请求参数，附加到URL 请求上
        params = {
            "extras[]": "count",
            "geohash": "wx4g0bmjetr7",
            "keyword": "%s" % keyword,
            "limit": "20",
            "type": "nearby"
        }
        params = urllib.parse.urlencode(params)  # 将请求参数进行编码
        req.full_url = newUrl + params  # 重新构造请求参数

        webpage = urllib.request.urlopen(req)  # 获取数据
        contentBytes = webpage.read().decode("utf-8")
        jsondata = json.loads(contentBytes)  # 将数据解析成json格式
        write2Excel(jsondata, keyword)  # 将数据写入excel 中
